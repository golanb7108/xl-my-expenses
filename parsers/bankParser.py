__author__ = 'gbenami'

"""
xl-my-expenses - A utility for parsing a bank sheet into
your expenses excel.
Bank Parser is an interface of different bank parsers.
"""

from openpyxl import load_workbook
import re

class bankParser(object):
    def __init__(self, fileName, bankName):
        """
        Opens the input file/stream and gets ready to parse it.
        """
        self.fileName = fileName
        self.bankName = bankName
        self.wb = load_workbook(filename = self.fileName, read_only=False)
        first_sheet = self.wb.get_sheet_names()[0]
        self.ws = self.wb.get_sheet_by_name(first_sheet)
        self.currentRow = 0
        self.currentRecord = 0
        self.wsRows = []
        self.recordsList = []
        self.datePatt = [re.compile("(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"),
                         re.compile("(\d{2})/(\d{2})/(\d{4})")]

        # Get all rows
        rowsToIter = "A1:G" + str(self.ws.get_highest_row())
        self.wsRows = [row for row in self.ws.iter_rows(rowsToIter)]

    def hasMoreRecords(self):
        """
        Checks if there is more records, and sets the current record
        to be the next record.
        :return: True if there are.
        """
        raise NotImplementedError( "bankParser is an abstract class" )

    def getRecord(self):
        """
        Get the current record.
        :return: the current record that is pointed.
        """
        return self.currentRecord

    def getDate(self):
        """
        Get the date of the expense.
        :return: return the date.
        """
        return self.currentRecord.getDate()

    def getCost(self):
        """
        Get the cost of the current record.
        :return: the cost.
        """
        return self.currentRecord.getCost()

    def getCollector(self):
        """
        Get the collector of the current record.
        :return: the Collector.
        """
        return self.currentRecord.getCosllector()

    def __str__(self):
        return "%s parser is parsing file %s." % (self.bankName, self.fileName)