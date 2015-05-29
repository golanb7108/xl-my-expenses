# -*- coding: utf-8 -*-
__author__ = 'gbenami'

"""
xl-my-expenses - A utility for parsing a bank sheet into
your expenses excel.
my Xl Writer is an instance of XL writer.
"""

from openpyxl import load_workbook
from xlWriter import xlWriter
from openpyxl import workbook
from constants import constants

class myXlWriter(xlWriter):
    def __init__(self, fileName, writerType):
        """
        Opens the input file/stream and gets ready to parse it.
        """
        xlWriter.__init__(self, fileName, writerType)
        self.emptyRowsPerSheet = {}

    def _findFreeRow(self, record):
        """
        Find free row to write records to.
        """
        rowsToIter = None
        expenseStartColumn = ''
        expenseStartRow = ''
        expenseEndColumn = ''
        expenseEndRow = ''
        sheetName = constants.MONTH_NUM_TO_STRING[record.getDate().tm_mon]
        # check if already found the the empty row
        if sheetName in self.emptyRowsPerSheet:
           self.emptyRowsPerSheet[sheetName] = (self.emptyRowsPerSheet[sheetName][0] + 1,
                                                self.emptyRowsPerSheet[sheetName][1])
        else:
            self.ws = self.wb.get_sheet_by_name(sheetName)
            recordsSignature = [constants.DATE_SIG, constants.CATEGORY_SIG,
                                constants.DESCRIPTION_SIG, constants.COST_SIG]

            for row in self.ws.iter_rows(row_offset=0):
                for cell in row:
                    if len(cell.column) > 1 or ord(cell.column) - 64 + 3 >= len(row):
                        break
                    rowTuple = [row[ord(cell.column) - 64 + i].value for i in range(4)]
                    if rowTuple == recordsSignature:
                        expenseStartColumn = str(cell.column)
                        expenseStartRow = str(cell.row)
                        expenseEndColumn = str(chr(ord(cell.column.lower()) + 3)).upper()
                        expenseEndRow = str(cell.row + 56)
                        rowsToIter = "%s%s:%s%s" % (expenseStartColumn, expenseStartRow,
                                                    expenseEndColumn, expenseEndRow)

            if rowsToIter == None:
                raise ValueError('The sheet does not have an expenses table.')

            wsRows = [row for row in self.ws.iter_rows(rowsToIter)]
            foundEmptyLine = False
            emptyRow = -1
            for rowNumber, line in enumerate(wsRows):
                currentRow = (line[i].value for i in range(4))
                if all(item is None for item in currentRow) and not foundEmptyLine:
                    foundEmptyLine = True
                    emptyRow = rowNumber
                elif not all(item is None for item in currentRow) and foundEmptyLine:
                    foundEmptyLine = False
            firstEmptyRow = emptyRow + int(expenseStartRow)

            self.emptyRowsPerSheet[sheetName] = (firstEmptyRow, (chr(ord(expenseStartColumn.lower()) + 1)).upper())

        return self.emptyRowsPerSheet[sheetName]

    def writeRecord(self, record):
        """
        Checks if there is more records, and sets the current record
        to be the next record.
        :return: True if there are.
        """
        (emptyRow, emptyColumn) = self._findFreeRow(record)
        dateSlot = "%s%s" % (str(emptyColumn), str(emptyRow))
        collectorSlot = "%s%s" % (str(chr(ord(emptyColumn.lower()) + 2)).upper(), str(emptyRow))
        costSlot = "%s%s" % (str(chr(ord(emptyColumn.lower()) + 3)).upper(), str(emptyRow))

        self.infoToSave[(self.ws.title, dateSlot)] = "%s/%s/%s" % (record.getDate().tm_mday,
                                                                   record.getDate().tm_mon,record.getDate().tm_year)
        self.infoToSave[(self.ws.title, collectorSlot)] = record.getCollector()
        self.infoToSave[(self.ws.title, costSlot)] = record.getCost()