__author__ = 'gbenami'

"""
xl-my-expenses - A utility for parsing a bank sheet into
your expenses excel.
xl Writer is an interface of different xl writers.
"""

from openpyxl import load_workbook
from openpyxl import workbook

class xlWriter(object):
    def __init__(self, fileName, writerType):
        """
        Opens the input file/stream and gets ready to parse it.
        """
        self.fileName = fileName
        self.writerType = writerType
        self.wb = load_workbook(filename = self.fileName, read_only=False)
        self.infoToSave = {}

    def _findFreeRow(self, record):
        """
        Find free row to write records to.
        """
        raise NotImplementedError( "xlWriter is an abstract class" )

    def writeRecord(self, record):
        """
        Checks if there is more records, and sets the current record
        to be the next record.
        :return: True if there are.
        """
        raise NotImplementedError( "xlWriter is an abstract class" )

    def saveFile(self, accountManageFileName):
        """
        Save the destination file.
        """
        from xlwings import Workbook, Range

        wb = Workbook(accountManageFileName)
        for placeToWrite in self.infoToSave:
            sheet = placeToWrite[0]
            cell = placeToWrite[1]
            data = self.infoToSave[placeToWrite]
            Range(sheet, cell).value = data
        wb.save(accountManageFileName)

    def __str__(self):
        return "%s writer is writing to file %s." % (self.writerType, self.fileName)
